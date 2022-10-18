#!/usr/bin/env python
# coding: utf-8

# %%
# !pip install keras

# %%

import csv

import os

import numpy as np

import tensorflow as tf

import keras

import copy

import glob

import openpyxl as excel

from skimage import data

from PIL import Image

from keras.preprocessing.image import array_to_img

from keras.utils.np_utils import to_categorical

from collections import OrderedDict

from pathlib import Path

from tqdm import tqdm

from natsort import natsorted

import pickle


from keras.models import Sequential

from keras.layers import Dense, Dropout, Flatten, Conv2D, MaxPooling2D, BatchNormalization

from keras.utils import np_utils

# from sklearn.datasets import fetch_mldata

import matplotlib

import matplotlib.pyplot as plt

from keras.models import Model

from keras.models import model_from_json

from keras.models import load_model

from tensorflow.keras.optimizers import RMSprop


# %%

# データセットを読み込んで学習データと訓練データに分ける
# そして保存するためのファイル

# クラス数
sub_num = 34

# 順番が保持された辞書
dict = OrderedDict({
    '2km': [],
    '3km': [],
    '4km': [],
    '5km': [],
    '6km': [],
    '7km': [],
    '8km': [],
    '9km': [],
    '10km': [],
})

# dictを異なるIDでコピーする(dictが変更されても影響は受けない)
gxdata = copy.deepcopy(dict)
gydata = copy.deepcopy(dict)
pxdata = []
pydata = []

for i in range(3):
    pxdata.append(copy.deepcopy(dict))
    pydata.append(copy.deepcopy(dict))

# naturalsort #["00001","00002"....]
sblist = natsorted(glob.glob("data/TreadmillDatasetA/*"))
for sbi, sb in enumerate(sblist):  # index object
    # リスト型 ["gallery_2km","gallery_3km"....]
    galist = natsorted(glob.glob(sb+"/gallery*"))
    prlist = natsorted(glob.glob(sb+"/probe*"))

    # gallery読み込み
    for kmi, km in enumerate(galist, 2):  # 2kmからだからindexの開始数値は2
        key = str(kmi)+'km'
        # ["00000001.png","00000002.png"....]
        piclist = natsorted(glob.glob(km+"/*.png"))
        for pic in piclist:
            tmp = np.array(Image.open(pic)).reshape(128, 88, 1)
            tmp = tmp/255  # 画像を正規化
            gxdata[key].append(tmp)
            gydata[key].append(sbi)  # label

    # probe読み込み
    # 3つにわける
    for kmi, km in enumerate(prlist, 2):
        key = str(kmi)+'km'
        piclist = natsorted(glob.glob(km+"/*.png"))
        for pici, pic in enumerate(piclist):
            tmp = np.array(Image.open(pic)).reshape(128, 88, 1)
            tmp = tmp/255
            if(pici < len(piclist) / 3):
                i = 0
            elif(pici < len(piclist) / 3*2):
                i = 1
            else:
                i = 2
            pxdata[i][key].append(tmp)
            pydata[i][key].append(sbi)  # label

# %%
print(gydata['2km'][500])

# %%


glabel = copy.deepcopy(gydata)
plabel = copy.deepcopy(pydata)


# %%


for km in gxdata.keys():  # ["2km","3km",....]
    gxdata[km] = np.array(gxdata[km])
    gxdata[km].astype('float32')
    gydata[km] = np.array(gydata[km])
    gydata[km] = to_categorical(gydata[km])

# %%


for i in range(3):
    for km in pxdata[i].keys():  # ["2km","3km",....]
        pxdata[i][km] = np.array(pxdata[i][km])
        pxdata[i][km].astype('float32')
        pydata[i][km] = np.array(pydata[i][km])
        pydata[i][km] = to_categorical(pydata[i][km])


# %%


# softmax

def smax(orgmodel, xdata, ydata, subnum):
    prob = 0
    sxmodel = Model(inputs=orgmodel.input, outputs=orgmodel.layers[len(
        orgmodel.layers)-1].output)  # 最終層の一つ前の層
    sxmatrix = sxmodel.predict(xdata)  # 最終層の一つ前の層まで予測
    for i in range(subnum):
        ff = ydata.index(i)  # ex. ydata = ["00001","00003","00008",....]
#         print(ff)
        if(i == (subnum-1)):
            lf = len(ydata)
        else:
            lf = ydata.index(i+1)
        res = sum(sxmatrix[:][ff:lf]).argmax()  # 一番大きいindexを返す [:]は全ての要素
        if(i == res):
            prob += 1
    prob = prob/subnum*100

    return prob


# %%


# モデル宣言

model = Sequential()

# レイヤー追加
model.add(Conv2D(18, kernel_size=(7, 7),
                 activation='relu',
                 input_shape=(128, 88, 1)))
model.add(MaxPooling2D(pool_size=(2, 2), strides=2))
model.add(BatchNormalization())
model.add(Conv2D(45, kernel_size=(5, 5),
                 activation='relu'))
model.add(MaxPooling2D(pool_size=(3, 3), strides=2))
model.add(BatchNormalization())
model.add(Flatten())
model.add(Dense(1024, activation='relu'))
model.add(Dense(sub_num, activation='softmax'))
model.compile(loss='categorical_crossentropy',
              optimizer='sgd', metrics=['accuracy'])

model.save('./path/to/model.h5')


# %%

epoch = 5

result = np.zeros((9, 9))  # 9×9の2次元配列生成
for tri, trkm in enumerate(gxdata.keys()):
    #     exmodel = copy.deepcopy(model)
    exmodel = load_model('path/to/model.h5')
    # verboseはログの出力の指定(2ならepochごとに1行のログを出力)
    exmodel.fit(gxdata[trkm], gydata[trkm],
                epochs=epoch, batch_size=10, verbose=2)
    outmodel = Model(inputs=exmodel.input, outputs=exmodel.layers[len(
        exmodel.layers)-2].output)  # fc層からのCNN特徴取り出しのモデル
    feature = outmodel.predict(gxdata[trkm], batch_size=10, verbose=0)

    if not os.path.exists('./cnnfeature/silhouette/cnnmodel/'+trkm):
        os.mkdir('./cnnfeature/silhouette/cnnmodel/'+trkm)  # 保存先のディレクトリ作成
    wb = excel.Workbook()
    wb.save(f'./cnnfeature/silhouette/cnnmodel/{trkm}/tr.csv')  # excelファイル作成
    # 作成したexcelファイルにgallery特徴量保存 .Tで転置
    np.savetxt(
        f'./cnnfeature/silhouette/cnnmodel/{trkm}/tr.csv', feature.T, delimiter=',')
    open(
        f'./cnnfeature/silhouette/cnnmodel/tr{trkm}.json', "w").write(exmodel.to_json())
    exmodel.save_weights(
        f'./cnnfeature/silhouette/cnnmodel/tr{trkm}.h5')  # 重み保存
    for n in range(3):
        for tsi, tskm in enumerate(pxdata[n].keys()):
            feature = outmodel.predict(
                pxdata[n][tskm], batch_size=10, verbose=0)
            wb = excel.Workbook()
            wb.save(
                f'./cnnfeature/silhouette/cnnmodel/{trkm}/ts{tskm}_{n+1}.csv')
            # probe特徴量保存 3つに分けたうちの1つずつ
            np.savetxt(
                f'./cnnfeature/silhouette/cnnmodel/{trkm}/ts{tskm}_{n+1}.csv', feature.T, delimiter=',')
            sxres = smax(exmodel, pxdata[n][tskm], plabel[n][tskm], sub_num)
            result[tri][tsi] += sxres
        print(result[tri])
result /= (n+1)  # 3つの平均を取るために3で割る
wb = excel.Workbook()
wb.save(f'./results/silhouette/softmax.csv')
np.savetxt(f'./results/silhouette/softmax.csv', result, delimiter=',')
