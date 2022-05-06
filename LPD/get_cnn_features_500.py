# %%
import csv

import os

import numpy as np

import tensorflow as tf

import pandas as pd

import keras

import copy

import glob

import openpyxl as excel

from skimage import data

from PIL import Image

from keras.preprocessing.image import array_to_img

from keras.utils.np_utils import to_categorical


from collections import OrderedDict

from pathlib import Path

from tqdm import tqdm

from natsort import natsorted

import pickle


from keras.models import Sequential

from keras.layers import Dense, Dropout, Flatten, Conv2D, MaxPooling2D, BatchNormalization

from keras.utils import np_utils

# from sklearn.datasets import fetch_mldata

import matplotlib

import matplotlib.pyplot as plt

from keras.models import Model

from keras.models import model_from_json

from keras.models import load_model

from tensorflow.keras.optimizers import RMSprop

# %%
# クラス数
sub_num = 500

# 前準備
# csvファイルの読み込み
# まとめて実行する場合
# df_gallery = pd.read_csv("./LPD/OULP-C1V2_Pack/OULP-C1V2_SubjectIDList(FormatVersion1.0)/IDList_OULP-C1V2-A-75_Gallery.csv", header=None)
# df_probe = pd.read_csv("./LPD/OULP-C1V2_Pack/OULP-C1V2_SubjectIDList(FormatVersion1.0)/IDList_OULP-C1V2-A-75_Probe.csv", header=None)

# セルごとに実行する場合
df_gallery = pd.read_csv(
    "./OULP-C1V2_Pack/OULP-C1V2_SubjectIDList(FormatVersion1.0)/IDList_OULP-C1V2-A-75_Gallery.csv", header=None)
df_probe = pd.read_csv(
    "./OULP-C1V2_Pack/OULP-C1V2_SubjectIDList(FormatVersion1.0)/IDList_OULP-C1V2-A-75_Probe.csv", header=None)

# id,画像の初め終わりの情報を入れた配列を作る
df_gallery_id = df_gallery.iloc[0:sub_num, 0]
df_probe_id = df_probe.iloc[0:sub_num, 0]
df_ga_start = df_gallery.iloc[0:sub_num, 2]
df_ga_end = df_gallery.iloc[0:sub_num, 3]
df_pr_start = df_probe.iloc[0:sub_num, 2]
df_pr_end = df_probe.iloc[0:sub_num, 3]

df_gallery_id = np.array(df_gallery_id)
df_probe_id = np.array(df_probe_id)
df_ga_start = np.array(df_ga_start)
df_ga_end = np.array(df_ga_end)
df_pr_start = np.array(df_pr_start)
df_pr_end = np.array(df_pr_end)

# 二次元配列を一次元配列に変換する
df_gallery_id = df_gallery_id.flatten()
df_probe_id = df_probe_id.flatten()
df_ga_start = df_ga_start.flatten()
df_ga_end = df_ga_end.flatten()
df_pr_start = df_pr_start.flatten()
df_pr_end = df_pr_end.flatten()

# 全ての数字を同じ桁に合わせる
gallery_id = []
probe_id = []
gallery_start = []
gallery_end = []
probe_start = []
probe_end = []

for i in range(sub_num):
    gallery_id.append(str(df_gallery_id[i]).zfill(7))
    probe_id.append(str(df_probe_id[i]).zfill(7))
    gallery_start.append(str(df_ga_start[i]).zfill(8))
    gallery_end.append(str(df_ga_end[i]).zfill(8))
    probe_start.append(str(df_pr_start[i]).zfill(8))
    probe_end.append(str(df_pr_end[i]).zfill(8))


# %%
g_list = []
p_list = []
# g_list = natsort(
#     glob.glob("OULP-C1V2_Pack/OULP-C1V2_NormalizedSilhouette(88x128)/Seq00/*"))
# p_list = []

for i in range(sub_num):
    g_list.append(glob.glob(
        f"OULP-C1V2_Pack/OULP-C1V2_NormalizedSilhouette(88x128)/Seq00/{gallery_id[i]}"))
    p_list.append(glob.glob(
        f"OULP-C1V2_Pack/OULP-C1V2_NormalizedSilhouette(88x128)/Seq01/{probe_id[i]}"))

# 一次元配列にする
g_list = np.array(g_list)
p_list = np.array(p_list)
g_list = g_list.flatten()
p_list = p_list.flatten()

print(gallery_start[0])
print(gallery_end[0])

# %%
gxdata = []
gydata = []

# %%
# gallery読み込み
for gai, ga in enumerate(g_list):
    pic_num = int(gallery_start[gai])
    ga_piclist = []
    while pic_num <= int(gallery_end[gai]):
        pic_list = glob.glob(ga+f"/{str(pic_num).zfill(8)}.png")
        ga_piclist.append(pic_list[0])
        pic_num += 1
    for pic in ga_piclist:
        tmp = np.array(Image.open(pic)).reshape(128, 88, 1)
        tmp = tmp/255  # 画像を正規化
        gxdata.append(tmp)
        gydata.append(gai)  # label

# %%
# probe読み込み
# 3つに分ける
pxdata = []
pydata = []
# for i in range(3):
#     array = []
#     pxdata.append(array)
#     pydata.append(array)

pxdata1 = []
pxdata2 = []
pxdata3 = []
pydata1 = []
pydata2 = []
pydata3 = []


# %%
for pri, pr in enumerate(p_list):
    pic_num = int(probe_start[pri])
    pr_piclist = []
    while pic_num <= int(probe_end[pri]):
        pic_list = glob.glob(pr+f"/{str(pic_num).zfill(8)}.png")
        pr_piclist.append(pic_list[0])
        pic_num += 1
    for pici, pic in enumerate(pr_piclist):
        tmp = np.array(Image.open(pic)).reshape(128, 88, 1)
        tmp = tmp/255
        if(pici < len(pr_piclist) / 3):
            pxdata1.append(tmp)
            pydata1.append(pri)
        elif(pici < len(pr_piclist) / 3*2):
            i = 1
            pxdata2.append(tmp)
            pydata2.append(pri)
        else:
            i = 2
            pxdata3.append(tmp)
            pydata3.append(pri)
        # print(tmp)
        # pxdata[i].append(tmp)
        # pydata[i].append(pri)

pxdata.append(pxdata1)
pxdata.append(pxdata2)
pxdata.append(pxdata3)

pydata.append(pydata1)
pydata.append(pydata2)
pydata.append(pydata3)

# %%
glabel = copy.deepcopy(gydata)
plabel = copy.deepcopy(pydata)

gxdata = np.array(gxdata)
gxdata.astype('float32')
gydata = np.array(gydata)
gydata = to_categorical(gydata, sub_num)

# %%
for i in range(3):
    pxdata[i] = np.array(pxdata[i])
    print(pxdata[i].shape)
    pxdata[i].astype('float32')
    pydata[i] = np.array(pydata[i])
    pydata[i] = to_categorical(pydata[i], sub_num)


# %%

# softmax

def smax(orgmodel, xdata, ydata, subnum):
    prob = 0
    sxmodel = Model(inputs=orgmodel.input, outputs=orgmodel.layers[len(
        orgmodel.layers)-1].output)  # 最終層の一つ前の層
    sxmatrix = sxmodel.predict(xdata)  # 最終層の一つ前の層まで予測
    for i in range(subnum):
        ff = ydata.index(i)
#         print(ff)
        if(i == (subnum-1)):
            lf = len(ydata)
        else:
            lf = ydata.index(i+1)
        res = sum(sxmatrix[:][ff:lf]).argmax()  # 一番大きいindexを返す [:]は全ての要素
        if(i == res):
            prob += 1
    prob = prob/subnum*100

    return prob


# %%
model = Sequential()

# レイヤー追加
model.add(Conv2D(18, kernel_size=(7, 7),
                 activation='relu',
                 input_shape=(128, 88, 1)))
model.add(MaxPooling2D(pool_size=(2, 2), strides=2))
model.add(BatchNormalization())
model.add(Conv2D(45, kernel_size=(5, 5),
                 activation='relu'))
model.add(MaxPooling2D(pool_size=(3, 3), strides=2))
model.add(BatchNormalization())
model.add(Flatten())
model.add(Dense(1024, activation='relu'))
model.add(Dense(sub_num, activation='softmax'))
model.compile(loss='categorical_crossentropy',
              optimizer='sgd', metrics=['accuracy'])

model.save('./path/to/model.h5')


# %%

epoch = 5


result = 0

exmodel = load_model('path/to/model.h5')
exmodel.fit(gxdata, gydata, epochs=epoch, batch_size=10, verbose=2)
outmodel = Model(inputs=exmodel.input, outputs=exmodel.layers[len(
    exmodel.layers)-2].output)  # fc層からのCNN特徴取り出しのモデル
feature = outmodel.predict(gxdata, batch_size=10, verbose=0)

wb = excel.Workbook()
wb.save(f'./cnnfeature/silhouette/cnnmodel/500.csv')  # excelファイル作成
# 作成したexcelファイルにgallery特徴量保存 .Tで転置
np.savetxt(f'./cnnfeature/silhouette/cnnmodel/500.csv',
           feature.T, delimiter=',')
open(f'./cnnfeature/silhouette/cnnmodel/500.json', "w").write(exmodel.to_json())
exmodel.save_weights(f'./cnnfeature/silhouette/cnnmodel/500.h5')  # 重み保存
for n in range(3):
    # for tsi, tskm in enumerate(pxdata[n].keys()):
    feature = outmodel.predict(pxdata[n], batch_size=10, verbose=0)
    wb = excel.Workbook()
    wb.save(f'./cnnfeature/silhouette/cnnmodel/500_{n+1}.csv')
    # probe特徴量保存 3つに分けたうちの1つずつ
    np.savetxt(
        f'./cnnfeature/silhouette/cnnmodel/500_{n+1}.csv', feature.T, delimiter=',')
    sxres = smax(exmodel, pxdata[n], plabel[n], sub_num)
    result += sxres
result /= (n+1)  # 3つの平均を取るために3で割る
print(result)
# wb = excel.Workbook()
# wb.save(f'./results/silhouette/softmax.csv')
# np.savetxt(f'./results/silhouette/softmax.csv', result, delimiter=',')



# %%
